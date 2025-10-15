from fastapi import APIRouter, UploadFile, File, Depends, HTTPException, status
from typing import List, Dict, Any
from sqlalchemy.ext.asyncio import AsyncSession
from db.session import get_db
from core.security import get_current_user
from models.user import User
from models.product import Product
import csv
import io
import tempfile
from sqlalchemy import select
import os

router = APIRouter()


def _iter_rows_from_csv_bytes(content: bytes, encoding: str = 'utf-8'):
    text = content.decode(encoding, errors='replace')
    reader = csv.DictReader(io.StringIO(text))
    for row in reader:
        yield row


def _rows_iterator_for_content(content: bytes, filename: str, content_type: str):
    """Return a generator for rows depending on file type"""
    ext = (filename or '').lower()
    if ext.endswith('.csv') or content_type in ['text/csv', 'application/csv']:
        return _iter_rows_from_csv_bytes(content)
    else:
        return _iter_rows_from_xlsx_bytes(content)


def _iter_rows_from_xlsx_bytes(content: bytes):
    # write to temp file and use openpyxl
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tf:
        tf.write(content)
        tmpname = tf.name
    try:
        try:
            from openpyxl import load_workbook
        except Exception as e:
            raise RuntimeError('openpyxl required to parse xlsx files') from e
        wb = load_workbook(tmpname, read_only=True)
        ws = wb.active
        # Normalize header values to strings (some headers may be numeric or None)
        first_row = next(ws.rows)
        headers = [str(cell.value).strip() if cell.value is not None else '' for cell in first_row]
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Build a dict mapping header->cell value
            out = {}
            for i in range(len(headers)):
                key = headers[i] if i < len(headers) else f'col_{i}'
                out[key] = row[i] if i < len(row) else None
            yield out
    finally:
        try:
            os.unlink(tmpname)
        except Exception:
            pass


def _normalize_row(row: Dict[str, Any]) -> Dict[str, Any]:
    # Accept case-insensitive headers: sku, name, cantidad, precio
    out = {}
    mapping = {}
    for k in row.keys():
        if k is None:
            continue
        kstr = str(k).strip().lower()
        mapping[kstr] = k

    def get(key):
        k = mapping.get(key)
        return row.get(k) if k is not None else None

    def to_str_safe(v):
        if v is None:
            return None
        try:
            return str(v).strip()
        except Exception:
            return None

    sku_val = get('sku')
    name_val = get('name') or get('nombre')
    category_val = get('category') or get('categoria')
    out['sku'] = to_str_safe(sku_val)
    out['name'] = to_str_safe(name_val)
    out['category'] = to_str_safe(category_val)
    # cantidad
    qty = get('cantidad') or get('quantity') or get('qty')
    qty_s = to_str_safe(qty)
    try:
        out['quantity'] = int(float(qty_s.replace(',', ''))) if qty_s is not None and qty_s != '' else None
    except Exception:
        out['quantity'] = None
    # price
    price = get('precio') or get('price')
    price_s = to_str_safe(price)
    try:
        out['price'] = float(price_s.replace(',', '')) if price_s is not None and price_s != '' else None
    except Exception:
        out['price'] = None

    return out


def _validate_row(row: Dict[str, Any]) -> List[str]:
    errors = []
    if not row.get('name'):
        errors.append('name missing')
    if not row.get('category'):
        errors.append('category missing')
    if row.get('quantity') is None:
        errors.append('quantity invalid')
    elif row.get('quantity') < 0:
        errors.append('quantity negative')
    if row.get('price') is None:
        errors.append('price invalid')
    elif row.get('price') < 0:
        errors.append('price negative')
    return errors


@router.post('/products/verify')
async def verify_products_import(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """Verify uploaded CSV/XLSX and return preview and errors. Does not persist."""
    # Size limit 20MB
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='File too large (max 20MB)')

    rows_iter = _rows_iterator_for_content(content, file.filename, file.content_type)

    preview = []
    errors_report = []
    total = 0
    critical = 0
    limit_rows = 200000
    for row in rows_iter:
        total += 1
        if total > limit_rows:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Too many rows in file; split file into smaller parts')
        nr = _normalize_row(row)
        errs = _validate_row(nr)
        is_critical = len([e for e in errs if 'missing' in e or 'invalid' in e or 'negative' in e]) > 0
        if is_critical:
            critical += 1
        preview.append({'row': total, 'data': nr, 'errors': errs})
        if len(preview) >= 1000:
            # limit preview size returned
            break

    return {
        'filename': file.filename,
        'total_rows': total,
        'preview_count': len(preview),
        'critical_errors': critical,
        'preview': preview,
    }


@router.post('/products')
async def import_products(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Perform import: insert/update products from CSV/XLSX. Returns summary."""
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='File too large (max 20MB)')

    # First, run a quick validation pass to detect critical errors and total rows
    rows_iter_for_check = _rows_iterator_for_content(content, file.filename, file.content_type)
    total_check = 0
    critical_check = 0
    limit_rows = 200000
    for row in rows_iter_for_check:
        total_check += 1
        if total_check > limit_rows:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Too many rows in file; split file')
        nr = _normalize_row(row)
        errs = _validate_row(nr)
        is_critical = len([e for e in errs if 'missing' in e or 'invalid' in e or 'negative' in e]) > 0
        if is_critical:
            critical_check += 1

    if critical_check > 0:
        # Block import if there are critical errors
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f'Import blocked: {critical_check} critical row errors detected. Please verify the file before importing.')

    # If check passed, create an iterator again for actual processing
    rows_iter = _rows_iterator_for_content(content, file.filename, file.content_type)

    inserted = 0
    updated = 0
    failed = 0
    processed = 0
    limit_rows = 200000
    for row in rows_iter:
        processed += 1
        if processed > limit_rows:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Too many rows in file; split file')
        nr = _normalize_row(row)
        errs = _validate_row(nr)
        if errs:
            failed += 1
            continue

        # if SKU provided, try to update by SKU, else by name exact
        if nr.get('sku'):
            result = await db.execute(select(Product).filter(Product.sku == nr['sku']))
            prod = result.scalar_one_or_none()
        else:
            result = await db.execute(select(Product).filter(Product.name == nr['name']))
            prod = result.scalar_one_or_none()

        if prod:
            # update stock and price
            prod.stock = nr['quantity']
            prod.price = nr['price']
            prod.name = nr['name']
            prod.category = nr.get('category')
            if nr.get('sku'):
                prod.sku = nr['sku']
            updated += 1
        else:
            new = Product(name=nr['name'], price=nr['price'], stock=nr['quantity'], sku=nr.get('sku'), category=nr.get('category') or 'General')
            db.add(new)
            inserted += 1

    await db.flush()

    # Create audit log entry if model exists
    try:
        from models.audit_log import AuditLog
        result_str = 'success' if failed == 0 else 'partial_failure' if inserted+updated>0 else 'failure'
        entry = AuditLog(actor_user_id=current_user.id, action='import_products', result=result_str, detail=f'file={file.filename} inserted={inserted} updated={updated} failed={failed}')
        db.add(entry)
        await db.flush()
    except Exception:
        # No audit log model or insertion failed - ignore but do not break import
        pass

    return {
        'filename': file.filename,
        'processed': processed,
        'inserted': inserted,
        'updated': updated,
        'failed': failed,
    }
